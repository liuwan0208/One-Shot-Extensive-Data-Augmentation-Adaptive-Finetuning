import os
from os.path import join
import nibabel as nib
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import matplotlib as mpl
import seaborn as sns
import warnings; warnings.filterwarnings(action='once')
from scipy import stats
import math
import openpyxl

from statsmodels.stats.multicomp import (pairwise_tukeyhsd, MultiComparison)
from statsmodels.formula.api import ols


bundles = ['AF_left', 'AF_right', 'ATR_left', 'ATR_right', 'CA', 'CC_1', 'CC_2', 'CC_3', 'CC_4', 'CC_5', 'CC_6',
           'CC_7', 'CG_left', 'CG_right', 'CST_left', 'CST_right', 'MLF_left', 'MLF_right', 'FPT_left',
           'FPT_right', 'FX_left', 'FX_right', 'ICP_left', 'ICP_right', 'IFO_left', 'IFO_right', 'ILF_left',
           'ILF_right', 'MCP', 'OR_left', 'OR_right', 'POPT_left', 'POPT_right', 'SCP_left', 'SCP_right',
           'SLF_I_left', 'SLF_I_right', 'SLF_II_left', 'SLF_II_right', 'SLF_III_left', 'SLF_III_right',
           'STR_left', 'STR_right', 'UF_left', 'UF_right', 'CC', 'T_PREF_left', 'T_PREF_right', 'T_PREM_left',
           'T_PREM_right', 'T_PREC_left', 'T_PREC_right', 'T_POSTC_left', 'T_POSTC_right', 'T_PAR_left',
           'T_PAR_right', 'T_OCC_left', 'T_OCC_right', 'ST_FO_left', 'ST_FO_right', 'ST_PREF_left',
           'ST_PREF_right', 'ST_PREM_left', 'ST_PREM_right', 'ST_PREC_left', 'ST_PREC_right', 'ST_POSTC_left',
           'ST_POSTC_right', 'ST_PAR_left', 'ST_PAR_right', 'ST_OCC_left', 'ST_OCC_right']
training_subject=["992774", "991267", "987983", "984472", "983773", "979984", "978578", "965771", "965367", "959574",
            "958976", "957974", "951457", "932554", "930449", "922854", "917255", "912447", "910241",
            "904044", "901442", "901139", "901038", "899885", "898176", "896879", "896778", "894673", "889579",
            "877269", "877168", "872764", "872158", "871964", "871762", "865363", "861456", "859671",
            "857263", "856766", "849971", "845458", "837964", "837560", "833249", "833148", "826454", "826353",
            "816653", "814649", "802844", "792766", "792564", "789373", "786569", "784565", "782561", "779370",
            "771354", "770352"]

def Tract_SampleNum():
    input_path = '/data3/wanliu/HCP_Wasser/HCP_for_training_COPY'
    wb=openpyxl.Workbook()
    ws=wb.active

    ### -------- calculate sample relative proportion for 20-60 training subjects---------------
    # output_path = '/data4/wanliu/Compare_results/TrainSampleNum_percent.xlsx'
    # ws.cell(row=1, column=1, value='TrainSubNum')
    # for i in range(72):
    #     ws.cell(row=i+2,column=1,value=bundles[i])
    # Num = [20,30,40,50,60]
    # for i in range(len(Num)):
    #     cal_samp_sum=np.zeros([72])
    #     for subject in training_subject[0:Num[i]]:
    #         label_path = join(input_path, subject, "bundle_masks_72.nii.gz")
    #         label = nib.load(label_path).get_fdata()
    #         cal_samp=np.sum(label, axis=(0,1,2))
    #         cal_samp_sum+=cal_samp
    #     ws.cell(row=1,column=i+2,value=Num[i])
    #     cal_mean = np.mean(cal_samp_sum)
    #     cal_samp_percent = cal_samp_sum/cal_mean
    #     for j in range(72):
    #         ws.cell(row=j+2,column=i+2,value=cal_samp_percent[j])


    ### --------- order the relative proportion for 60 training subjects--------------------------
    output_path = '/data4/wanliu/Compare_results/TrainSample_percent_order.xlsx'
    cal_samp_sum = np.zeros([72])
    for subject in training_subject[0:60]:
        label_path = join(input_path, subject, "bundle_masks_72.nii.gz")
        label = nib.load(label_path).get_fdata()
        cal_samp = np.sum(label, axis=(0, 1, 2))
        cal_samp_sum += cal_samp
        print('done')
    cal_samp_percent = cal_samp_sum / np.mean(cal_samp_sum)
    samp_percent_list=[]
    for i in range(72):
        samp_percent_list.append(cal_samp_percent[i])
    sorted_id = sorted(range(72), key=lambda k:samp_percent_list[k])

    print('-----Done-----')
    for i in range(72):
        print('{}:{}-{}, percent:{}'.format(i+1, sorted_id[i], bundles[sorted_id[i]], samp_percent_list[sorted_id[i]]))
        ws.cell(row=i+2,column=1,value=i+1)
        ws.cell(row=i+2,column=2,value=bundles[sorted_id[i]])
        ws.cell(row=i+2,column=3,value=sorted_id[i])
        ws.cell(row=i+2,column=4,value=samp_percent_list[sorted_id[i]])
    wb.save(output_path)

    samp_percent_lcopy = samp_percent_list
    sorted_id_copy = sorted_id
    samp_percent_lcopy.sort(reverse=True)
    sorted_id_copy.reverse() ##used to draw axis
    InverseOrder_bundle = []
    for i in range(72):
        InverseOrder_bundle.append(bundles[sorted_id_copy[i]])

    figsize = 20, 10
    figure, ax = plt.subplots(figsize=figsize)
    plt.bar(range(72),samp_percent_lcopy,color='blue',width=.3, tick_label=InverseOrder_bundle)
    plt.ylim(0,8)
    plt.tick_params(labelsize=10)
    plt.tick_params(axis='x', rotation=90)
    plt.xlim((-1,72))
    ## remove the black short line in the axes!
    plt.tick_params(bottom=False, left=True, right=False, top=False,color='b')


    font2 = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 13,
             }
    plt.ylabel('Relative Proportion of Sample Number', font2)
    plt.xlabel('White Matter Tract', font2)
    # plt.show()
    ## set linewidth of axes
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')

    plt.savefig('/data4/wanliu/Compare_results/TrainSample_percent_order.png', dpi=200, bbox_inches='tight')

def make_dir(directory):
    if not os.path.exists(directory):
        os.makedirs(directory)


def cohen_d(x,y,z,w):
    return (abs((x-y)/math.sqrt((z**2+w**2)/2.0)))


def writetoxlsx(save_path, column, row, value):
    outwb = openpyxl.Workbook()  # open a writing file
    outws = outwb.create_sheet()  # creat sheet in this file
    outws.cell(column=column, row=row, value=value)
    outwb.save(save_path)


def calculate_tract_p_d():
    save_path = '//data4/wanliu/Compare_results/d-p/test/dp_value_2.5.xlsx'

    base_dice = np.load('/data4/wanliu/Compare_results/2.5mm_unet-deepsup/Dice_TractSegDeep_train60.npy')  # (20,72)
    prop_dice = np.load('/data4/wanliu/Compare_results/2.5mm_unet-deepsup/Dice_TractSegDeepAdplw_train60_embed36.npy')

    # base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/Dice_TractSegDeep_train60.npy')  # (20,72)
    # prop_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/Dice_TractSegDeepAdplw_train60_embed36.npy')

    # base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet/Dice_TractsegNodeep_train60.npy')  # (20,72)
    # prop_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet/Dice_TractSegAdplw_train60_embed36.npy')

    outwb = openpyxl.Workbook()
    outws=outwb.active

    ## tract-wise p
    p_list=[]

    # num= base_dice.shape[1]#tract
    num= base_dice.shape[0]#subject
    for i in range(num):
        # base_dice_tract = base_dice[:,i]#tract
        # prop_dice_tract = prop_dice[:,i]

        base_dice_tract = base_dice[i,:]#sub
        prop_dice_tract = prop_dice[i,:]


        ## p value---the paired student's t-test
        tract_ttest = stats.ttest_rel(base_dice_tract, prop_dice_tract)
        p = tract_ttest.pvalue
        p_list.append(p)

        ## effect sizes (cohen's d value---reflect the mean variation)
        base_tract_mean = np.mean(base_dice_tract)
        base_tract_std = np.std(base_dice_tract)
        prop_tract_mean = np.mean(prop_dice_tract)
        prop_tract_std = np.std(prop_dice_tract)
        es_cohen = cohen_d(base_tract_mean, prop_tract_mean, base_tract_std, prop_tract_std)

        outws.cell(column=1, row=i+7, value=bundles[i])
        outws.cell(column=2, row=i+7, value=p)
        outws.cell(column=6, row=i+7, value=es_cohen)
    sorted_id = sorted(range(len(p_list)), key=lambda k:p_list[k])
    p_k_list = list(range(num))
    for i in range(len(p_list)):
        p_k_list[sorted_id[i]] = i+1

    for i in range(num):
        p = p_list[i]
        k = p_k_list[i]
        Bonf_p_correction = p*num
        Holm_p_correction = p*(num-k+1)
        BH_p_correction = p*(num/k)
        outws.cell(column=3, row=i + 7, value=Bonf_p_correction)
        outws.cell(column=4, row=i + 7, value=Holm_p_correction)
        outws.cell(column=5, row=i + 7, value=BH_p_correction)

    outws.cell(column=1, row=6, value='bundle')
    outws.cell(column=2, row=6, value='p')
    outws.cell(column=3, row=6, value='p_bonf')
    outws.cell(column=4, row=6, value='p_holm')
    outws.cell(column=5, row=6, value='p_bh')
    outws.cell(column=6, row=6, value='d')

    outwb.save(save_path)


def calculate_one_p():

    # base_dice = np.load('/data4/wanliu/Compare_results/2.5mm_unet-deepsup/Dice_TractSegDeep_train60.npy')  # (20,72)
    # prop_dice = np.load('/data4/wanliu/Compare_results/2.5mm_unet-deepsup/Dice_TractSegDeepAdplw_train60_embed36.npy')

    # base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/Dice_TractSegDeep_train60.npy')  # (20,72)
    # prop_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/Dice_TractSegDeepAdplw_train60_embed36.npy')

    base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet/Dice_TractsegNodeep_train60.npy')  # (20,72)
    prop_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet/Dice_TractSegAdplw_train60_embed36.npy')

    base_dice_mean = np.mean(base_dice,axis=0) # 72_dim
    prop_dice_mean = np.mean(prop_dice,axis=0)
    tract_ttest = stats.ttest_rel(base_dice_mean, prop_dice_mean)
    print(base_dice_mean)
    print(prop_dice_mean)
    p = tract_ttest.pvalue
    print('p_value_72: ', p)

    base_dice_mean = np.mean(base_dice, axis=1)  # 20_dim
    prop_dice_mean = np.mean(prop_dice, axis=1)
    tract_ttest = stats.ttest_rel(base_dice_mean, prop_dice_mean)
    p = tract_ttest.pvalue
    print('p_value_20: ', p)


## combine segmentation bianry mask with FA
def create_result_image():
    FA_path = '/data3/wanliu/HCP_Wasser/HCP_for_training_COPY'
    label_path = '/data3/wanliu/HCP_Wasser/HCP_for_training_COPY'
    baselseg_path = '/data3/wanliu/TractSeg_Run/hcp_exp/my_custom_experiment_x9/segmentation_ep218/all_bund_seg'
    proposeg_path = '/data4/wanliu/TractSeg_Run/hcp_exp/my_custom_experiment/segmentation_ep898/all_bund_seg'
    save_path = '/data4/wanliu/Compare_results/FA_add_seg'

    test_subject = ["695768", "690152", "687163", "685058", "683256", "680957", "679568", "677968",
                "673455", "672756", "665254", "654754", "645551", "644044", "638049", "627549", "623844", "622236",
                "620434", "613538"]
    #test_subject = ["690152","687163", "685058", "683256"]

    for subject in test_subject:
        make_dir(join(save_path, subject))
        FA_nii = nib.load(join(FA_path, subject, 'FA.nii.gz'))
        FA_affine = FA_nii.affine
        FA = FA_nii.get_fdata()
        FA_copy =  np.transpose(np.array(np.dsplit(np.tile(FA,(1,1,72)),72)),[1,2,3,0])


        label = nib.load(join(label_path, subject, 'bundle_masks_72.nii.gz')).get_fdata()
        FA_label = FA_copy
        FA_label[label == 1] = 1
        nib.save(nib.Nifti1Image(FA_label.astype(np.float32), FA_affine), join(save_path, subject, 'FA_clabel.nii.gz'))

        baselseg = nib.load(join(baselseg_path, subject + '.nii.gz')).get_fdata()
        FA_baselseg = FA_copy
        FA_baselseg[baselseg == 1] = 1
        nib.save(nib.Nifti1Image(FA_baselseg.astype(np.float32), FA_affine), join(save_path, subject, 'FA_baselseg.nii.gz'))

        proposeg = nib.load(join(proposeg_path, subject + '.nii.gz')).get_fdata()
        FA_proposeg = FA_copy
        FA_proposeg[proposeg == 1] = 1
        nib.save(nib.Nifti1Image(FA_proposeg.astype(np.float32), FA_affine), join(save_path, subject, 'FA_proposeg.nii.gz'))

        print('save', subject)


## draw tract_wise dice comparison
def draw_tract_dice():
    ## load the mean dice files of different method
    base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/tractseg_attention/Dice_TractSegDeep_train60_selfB1.npy')
    # base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/Dice_TractSegDeep_train60.npy')#(20,72)
    # base_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/Dice_TractSegDeepAdplw_train60_embed36.npy')#(20,72)
    base_sub_mean = np.mean(base_dice, 1)  # 20
    base_tract_mean = np.mean(base_dice, 0)  # 72

    prop_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/tractseg_embed_attention/Dice_TractSegDeep_embed36_train60_selfb1.npy')
    # prop_dice = np.load('/data4/wanliu/Compare_results/1.25mm_unet-deepsup/tractseg_embed_fix_TW/Dice_TractSegDeep_embed36_train60_fix3w5.npy')
    prop_sub_mean = np.mean(prop_dice, 1)  # 20
    prop_tract_mean = np.mean(prop_dice, 0)  # 72

    save_file = '/data4/wanliu/Compare_results/1.25mm_unet-deepsup/tractseg_embed_attention/tract_embed36_train60_selfb1_comparewithTractSegDeepSelfB1.png'

    #prop_dice1 = np.load('/data4/wanliu/Compare_results/Dice_TractSegAdplw_train60_embed36_249lwrec0.npy')
    #prop_sub_mean1 = np.mean(prop_dice1, 1)  # 20
    #prop_tract_mean1 = np.mean(prop_dice1, 0)  # 72


    figsize = 10, 15
    figure, ax = plt.subplots(figsize=figsize)

    lineone, = plt.plot(base_tract_mean, bundles, 'gD', label='Baseline',markersize=5)
    #lineone, = plt.plot(base_tract_mean, bundles, 'g*', label='Baseline',markersize=5)

    #linethree, = plt.plot(prop_tract_mean1, bundles, 'bo', label='Baseline + Input Embed',markersize=5)
    linetwo, = plt.plot(prop_tract_mean, bundles, 'ro', label='Proposed',markersize=5)



    # lineone, = plt.plot(base_tract_mean, bundles, 'kx',label='Baseline', markersize=6,linewidth=15.0, alpha=0.7)
    # linetwo, = plt.plot(prop_tract_mean, bundles, 'k*',label='The proposed', markersize=8,linewidth=15.0, alpha=0.7)

    font1 = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 13,
             }
    # plt.legend(handles=[lineone, linetwo], prop=font1, frameon=False, facecolor='none', loc='upper left')
    plt.legend(handles=[lineone, linetwo], prop=font1, loc='upper left')
    #plt.legend(handles=[lineone, linethree, linetwo], prop=font1, loc='upper left')


    # plt.tick_params(labelsize=10)
    # plt.tick_params(axis='x', labelsize=9)
    plt.tick_params(labelsize=10, rotation=30)
    labels = ax.get_xticklabels() + ax.get_yticklabels()
    [label.set_fontname('Times New Roman') for label in labels]
    ## 1.25mm
    plt.xlim((0.65,0.95))
    plt.xticks(np.arange(0.65,0.95,0.025))
    ## 2.5mm
    # plt.xlim((0.525,0.9))
    # plt.xticks(np.arange(0.525,0.9,0.025))

    plt.ylim((-1,72))


    font2 = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 13,
             }
    plt.xlabel('Dice Coefficient', font2)
    plt.ylabel('White Matter Tract', font2)
    plt.grid(linestyle='--',linewidth=1, color='k',alpha=0.3, which='major')

    ## remove the black short line in the axes!
    plt.tick_params(bottom=False, left=False, right=False, top=False)

    ## set linewidth of axes
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')

    # plt.margins(0,0)
    plt.savefig(save_file, dpi=200, bbox_inches='tight')
    # plt.savefig('/data3/wanliu/Compare_results/figure_tract_warpf1_0.1.png', dpi=200, bbox_inches='tight')
    # plt.show()


def draw_sub_dice():
    ## load the mean dice files of different method
    base_dice = np.load('/data3/wanliu/Compare_results/Dice_218.npy')#(20,72)
    base_sub_mean = np.mean(base_dice, 1)  # 20
    base_tract_mean = np.mean(base_dice, 0)  # 72

    prop_dice = np.load('/data3/wanliu/Compare_results/Dice_932_twostage.npy')
    prop_sub_mean = np.mean(prop_dice, 1)  # 20
    prop_tract_mean = np.mean(prop_dice, 0)  # 72

    figure_size = 7,4
    fig, ax = plt.subplots(figsize=figure_size,)

    len=prop_sub_mean.shape
    base_y_list = np.ones(len)*0.5
    prob_y_list = np.ones(len)*1.0
    # lineone, = plt.plot(base_sub_mean, base_y_list+np.random.normal(0,0.02, size=len), 'o', label='Baseline (per subject)', alpha=0.7)
    # linetwo, = plt.plot(np.mean(base_sub_mean), base_y_list[0], 'b^', label='Baseline (mean)',markersize=8, alpha=0.8)
    # linethree, = plt.plot(prop_sub_mean, prob_y_list+np.random.normal(0,0.02, size=len), 'o', label='Proposed (per subject)', alpha=0.5)
    # linefour, = plt.plot(np.mean(prop_sub_mean), prob_y_list[0], 'r^', label='Proposed(mean)',markersize=8, alpha=0.8)

    lineone, = plt.plot(base_sub_mean, base_y_list + np.random.normal(0, 0.025, size=len), 'ko',
                        label='Baseline (per subject)', alpha=0.2)
    linetwo, = plt.plot(np.mean(base_sub_mean), base_y_list[0], 'k^', label='Baseline (mean)', markersize=8, alpha=0.5)
    linethree, = plt.plot(prop_sub_mean, prob_y_list + np.random.normal(0, 0.025, size=len), 'ko',
                          label='Proposed (per subject)', alpha=0.2)
    linefour, = plt.plot(np.mean(prop_sub_mean), prob_y_list[0], 'k^', label='Proposed(mean)', markersize=8, alpha=0.5)

    font = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 8,
             }
    # plt.legend(handles=[lineone, linetwo, linethree, linefour], prop=font, frameon=False, loc='upper left')

    plt.tick_params(axis='x',labelsize=8)
    plt.tick_params(axis='y',labelsize=8)

    plt.xticks(np.arange(0.815, 0.860, 0.005))

    y_labels=['','Baseline','Proposed','']
    plt.yticks(np.arange(0, 1.6, 0.5), y_labels)
    # ax.set_yticklabels(y_labels)


    ## remove the black short line in the axes!
    plt.tick_params(bottom=False, left=False, right=False, top=False)


    plt.grid(linestyle='dashdot')
    plt.xlabel('Dice Score', font)
    plt.ylabel('Method', font)

    ## set linewidth of axes
    ax.spines['bottom'].set_linewidth(1.5)
    ax.spines['top'].set_linewidth(1.5)
    ax.spines['left'].set_linewidth(1.5)
    ax.spines['right'].set_linewidth(1.5)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')


    plt.savefig('/data3/wanliu/Compare_results/figure_subject.png', dpi=400, bbox_inches='tight')


def draw_tract_dice_boxplot():
    ## load the mean dice files of different method
    base_dice = np.load('/data3/wanliu/Compare_results/Dice_218.npy')#(20,72)
    base_sub_mean = np.mean(base_dice, 1)  # 20
    base_tract_mean = np.mean(base_dice, 0)  # 72

    prop_dice = np.load('/data3/wanliu/Compare_results/Dice_932_twostage.npy')
    prop_sub_mean = np.mean(prop_dice, 1)  # 20
    prop_tract_mean = np.mean(prop_dice, 0)  # 72

    figure_size = 20,25
    fig, ax = plt.subplots(figsize=figure_size,)
    word_size=14
    box_linewidth=2
    box_width=0.4
    color_boxline='0.3'
    ax = plt.subplot(311)

    f1=plt.boxplot(base_dice[:, 0:24], positions=range(1,48,2),widths=box_width, showmeans=False, patch_artist=True,\
                boxprops={'color': color_boxline, 'facecolor':'turquoise','linewidth': box_linewidth*0.5}, \
                capprops={'color': color_boxline, 'linewidth': box_linewidth*0.7}, \
                whiskerprops={'color': color_boxline, 'linewidth': box_linewidth*0.7}, \
                medianprops={'color': color_boxline, 'linewidth': 0.5 * box_linewidth}, \
                flierprops={'marker': 'D', 'markersize': 4,'color':'0.5'}, \
                meanprops={'marker': '^', 'color': 'green', 'markersize': 5})
    f2 = plt.boxplot(prop_dice[:, 0:24], positions=range(2,49,2), widths=box_width, showmeans=False, patch_artist=True,\
                    boxprops={'color': color_boxline, 'facecolor': 'lightcoral', 'linewidth': box_linewidth * 0.5,}, \
                    capprops={'color': color_boxline, 'linewidth': box_linewidth * 0.7}, \
                    whiskerprops={'color': color_boxline, 'linewidth': box_linewidth * 0.7}, \
                    medianprops={'color': color_boxline, 'linewidth': 0.5 * box_linewidth}, \
                    flierprops={'marker': 'o', 'markersize': 4, }, \
                    meanprops={'marker': '^', 'color': 'green', 'markersize': 5})
    plt.legend([f1['boxes'][0],f2['boxes'][0]],['Baseline','Proposed'], loc='upper center', bbox_to_anchor=[0.5,1.1], fontsize=word_size,frameon=False, ncol=2)
    ## set cycle_color for one boxplot
    # color=['r','g',...]
    # for box, fliers, c in zip(f['boxes'],f['fliers'],colors):
    #     box.set(facecolor=c)
    #     fliers.set(markerfacecolor=c)
    plt.tick_params(labelsize=word_size, color='0.7', width=1)
    plt.xticks(np.arange(1.5, 48.5, 2), bundles[0:24], rotation=30)
    plt.yticks(np.arange(0.2, 1.1, 0.1))
    plt.ylabel('Dice Score', size=word_size)

    plt.grid(np.arange(3.5, 26, 3), 'major',color='0.7',linestyle='dashdot', axis='y')

    ax.spines['bottom'].set_linewidth(box_linewidth)
    ax.spines['top'].set_linewidth(box_linewidth)
    ax.spines['left'].set_linewidth(box_linewidth)
    ax.spines['right'].set_linewidth(box_linewidth)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')




    ax = plt.subplot(312)

    f1 = plt.boxplot(base_dice[:, 24:48], positions=range(1, 48, 2), widths=box_width, showmeans=False, patch_artist=True,\
                boxprops={'color': color_boxline, 'facecolor':'turquoise','linewidth': box_linewidth*0.5}, \
                capprops={'color': color_boxline, 'linewidth': box_linewidth*0.7}, \
                whiskerprops={'color': color_boxline, 'linewidth': box_linewidth*0.7}, \
                medianprops={'color': color_boxline, 'linewidth': 0.5 * box_linewidth}, \
                flierprops={'marker': 'D', 'markersize': 4,'color':'0.5'}, \
                meanprops={'marker': '^', 'color': 'green', 'markersize': 5})
    f2 = plt.boxplot(prop_dice[:, 24:48], positions=range(2, 49, 2),widths=box_width, showmeans=False, patch_artist=True,\
                    boxprops={'color': color_boxline, 'facecolor': 'lightcoral', 'linewidth': box_linewidth * 0.5,}, \
                    capprops={'color': color_boxline, 'linewidth': box_linewidth * 0.7}, \
                    whiskerprops={'color': color_boxline, 'linewidth': box_linewidth * 0.7}, \
                    medianprops={'color': color_boxline, 'linewidth': 0.5 * box_linewidth}, \
                    flierprops={'marker': 'o', 'markersize': 4, }, \
                    meanprops={'marker': '^', 'color': 'green', 'markersize': 5})

    plt.tick_params(labelsize=word_size, color='0.7', width=1)
    plt.xticks(np.arange(1.5, 48.5, 2), bundles[24:48], rotation=30)
    plt.yticks(np.arange(0.4, 1.1, 0.1))
    plt.ylabel('Dice Score', size=word_size)

    plt.grid(np.arange(3.5, 26, 3), 'major', color='0.7', linestyle='dashdot', axis='y')

    ax.spines['bottom'].set_linewidth(box_linewidth)
    ax.spines['top'].set_linewidth(box_linewidth)
    ax.spines['left'].set_linewidth(box_linewidth)
    ax.spines['right'].set_linewidth(box_linewidth)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')




    ax = plt.subplot(313)

    f1 = plt.boxplot(base_dice[:, 48:72], positions=range(1, 48, 2), widths=box_width, showmeans=False, patch_artist=True,\
                boxprops={'color': color_boxline, 'facecolor':'turquoise','linewidth': box_linewidth*0.5}, \
                capprops={'color': color_boxline, 'linewidth': box_linewidth*0.7}, \
                whiskerprops={'color': color_boxline, 'linewidth': box_linewidth*0.7}, \
                medianprops={'color': color_boxline, 'linewidth': 0.5 * box_linewidth}, \
                flierprops={'marker': 'D', 'markersize': 4,'color':'0.5'}, \
                meanprops={'marker': '^', 'color': 'green', 'markersize': 5})
    f2 = plt.boxplot(prop_dice[:, 48:72], positions=range(2, 49, 2), widths=box_width, showmeans=False, patch_artist=True,\
                    boxprops={'color': color_boxline, 'facecolor': 'lightcoral', 'linewidth': box_linewidth * 0.5,}, \
                    capprops={'color': color_boxline, 'linewidth': box_linewidth * 0.7}, \
                    whiskerprops={'color': color_boxline, 'linewidth': box_linewidth * 0.7}, \
                    medianprops={'color': color_boxline, 'linewidth': 0.5 * box_linewidth}, \
                    flierprops={'marker': 'o', 'markersize': 4, }, \
                    meanprops={'marker': '^', 'color': 'green', 'markersize': 5})

    plt.tick_params(labelsize=word_size, color='0.7', width=1)
    plt.xticks(np.arange(1.5, 48.5, 2), bundles[48:72], rotation=30)
    plt.yticks(np.arange(0.2, 1.1, 0.1))
    plt.ylabel('Dice Score', size=word_size)

    plt.grid(np.arange(3.5, 26, 3), 'major', color='0.7', linestyle='dashdot', axis='y')

    ax.spines['bottom'].set_linewidth(box_linewidth)
    ax.spines['top'].set_linewidth(box_linewidth)
    ax.spines['left'].set_linewidth(box_linewidth)
    ax.spines['right'].set_linewidth(box_linewidth)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')


    plt.subplots_adjust(hspace=0.16)

    # plt.show()
    plt.savefig('/data3/wanliu/Compare_results/figure_tract_boxplot.png', dpi=200,bbox_inches='tight')


def draw_tract_dice_boxplot1():
    ## load the mean dice files of different method
    base_dice = np.load('/data3/wanliu/Compare_results/Dice_218.npy')#(20,72)
    base_sub_mean = np.mean(base_dice, 1)  # 20
    base_tract_mean = np.mean(base_dice, 0)  # 72

    prop_dice = np.load('/data3/wanliu/Compare_results/Dice_932_twostage.npy')
    prop_sub_mean = np.mean(prop_dice, 1)  # 20
    prop_tract_mean = np.mean(prop_dice, 0)  # 72
    arr=np.transpose(np.concatenate([np.expand_dims(base_dice,-1),np.expand_dims(prop_dice,-1)],-1),[1,0,2])#72,20,2

    figure_size = 12,15
    figure, axes=plt.subplots(figsize=figure_size,)
    word_size=10
    box_linewidth=1
    axes_linewidth=1.5
    fliersize=3
    box_width=0.5
    whis=2

    ax = plt.subplot(311)
    cdf=pd.DataFrame()
    for i in range(24):
        df = pd.DataFrame(arr[i, :, :], columns=list(['Baseline','Proposed'])).assign(Trial=bundles[i])
        cdf= pd.concat([cdf, df])

    # plt.grid('major', color='0.7', linestyle='dashdot',axis='x')

    mdf = pd.melt(cdf, id_vars=['Trial'], var_name=['Method'])
    sns.boxplot(x="Trial", y="value", hue="Method", data=mdf, saturation=1, width=box_width,notch=True, whis=whis,\
                      linewidth=box_linewidth,palette={'Baseline':'orange','Proposed':'c'},fliersize=fliersize)  # RUN PLOT

    plt.tick_params(axis='y',labelsize=word_size, color='0.7', width=1)
    plt.tick_params(axis='x',labelsize=word_size, color='0.7', width=1,rotation=30)
    plt.yticks(np.arange(0.2, 1.1, 0.1))
    plt.ylabel('Dice Score', size=word_size)
    plt.xlabel('', size=word_size)
    plt.grid('major', color='0.7', linestyle='dashdot',axis='x')

    ax.legend(handles=None, labels=None)

    ax.spines['bottom'].set_linewidth(axes_linewidth)
    ax.spines['top'].set_linewidth(axes_linewidth)
    ax.spines['left'].set_linewidth(axes_linewidth)
    ax.spines['right'].set_linewidth(axes_linewidth)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')


    ax = plt.subplot(312)
    cdf = pd.DataFrame()
    for i in range(24,48):
        df = pd.DataFrame(arr[i, :, :], columns=list(['Baseline', 'Proposed'])).assign(Trial=bundles[i])
        cdf = pd.concat([cdf, df])

    mdf = pd.melt(cdf, id_vars=['Trial'], var_name=['Method'])
    sns.boxplot(x="Trial", y="value", hue="Method",data=mdf, palette={'Baseline':'orange','Proposed':'c'},whis=whis,\
                saturation=1, width=box_width, linewidth=box_linewidth,notch=True,
                fliersize=fliersize)  # RUN PLOT

    plt.tick_params(axis='y', labelsize=word_size, color='0.7', width=1)
    plt.tick_params(axis='x', labelsize=word_size, color='0.7', width=1, rotation=30)
    plt.yticks(np.arange(0.2, 1.1, 0.1))
    plt.ylabel('Dice Score', size=word_size)
    plt.xlabel('', size=word_size)
    plt.grid('major', color='0.7', linestyle='dashdot',axis='x')

    ax.legend(handles=None, labels=None)

    ax.spines['bottom'].set_linewidth(axes_linewidth)
    ax.spines['top'].set_linewidth(axes_linewidth)
    ax.spines['left'].set_linewidth(axes_linewidth)
    ax.spines['right'].set_linewidth(axes_linewidth)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')


    ax = plt.subplot(313)
    cdf = pd.DataFrame()
    for i in range(48,72):
        df = pd.DataFrame(arr[i, :, :], columns=list(['Baseline', 'Proposed'])).assign(Trial=bundles[i])
        cdf = pd.concat([cdf, df])

    mdf = pd.melt(cdf, id_vars=['Trial'], var_name=['Method'])
    sns.boxplot(x="Trial", y="value",hue="Method",data=mdf, palette={'Baseline':'orange','Proposed':'c'},whis=whis,\
                saturation=1, width=box_width, linewidth=box_linewidth,notch=True,
                fliersize=fliersize,)  # RUN PLOT

    plt.tick_params(axis='y', labelsize=word_size, color='0.7', width=1)
    plt.tick_params(axis='x', labelsize=word_size, color='0.7', width=1, rotation=30)
    plt.yticks(np.arange(0.2, 1.1, 0.1))
    plt.ylabel('Dice Score', size=word_size)
    plt.xlabel('', size=word_size)
    plt.grid('major', color='0.7', linestyle='dashdot',axis='x')

    ax.legend(handles=None, labels=None)

    ax.spines['bottom'].set_linewidth(axes_linewidth)
    ax.spines['top'].set_linewidth(axes_linewidth)
    ax.spines['left'].set_linewidth(axes_linewidth)
    ax.spines['right'].set_linewidth(axes_linewidth)
    ax.spines['bottom'].set_color('0.7')
    ax.spines['top'].set_color('0.7')
    ax.spines['left'].set_color('0.7')
    ax.spines['right'].set_color('0.7')



    plt.subplots_adjust(hspace=0.2)

    # adjust_box_widths(figure, 0.9)

    plt.savefig('/data3/wanliu/Compare_results/figure_tract_boxplot1.png', dpi=400,bbox_inches='tight')



if __name__ == "__main__":
    #Tract_SampleNum()
    # calculate_tract_p_d()
    # calculate_one_p()
    draw_tract_dice()
    #create_result_image()
    # draw_tract_dice_boxplot()
    # draw_tract_dice_boxplot1()


